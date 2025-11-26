import tempfile
from pathlib import Path
from typing import List
from io import BytesIO
from PIL import Image
import openpyxl
from openpyxl.chart import (
    BarChart, BarChart3D,
    LineChart, LineChart3D,
    PieChart, PieChart3D,
    AreaChart, AreaChart3D,
    ScatterChart,
    BubbleChart,
    RadarChart,
    DoughnutChart,
    StockChart,
    SurfaceChart, SurfaceChart3D
)
from openpyxl.utils import get_column_letter


class ExcelParser(BaseParser):
    def __init__(self) -> None:
        super().__init__(name="EXCEL_PARSER")
    
    def _get_used_range(self, worksheet):
        """
        Determine the actual used range of a worksheet, ignoring completely blank rows/columns.
        Returns (min_row, max_row, min_col, max_col)
        """
        # Get dimensions
        if worksheet.calculate_dimension() == 'A1:A1' and worksheet['A1'].value is None:
            return None
        
        min_row = worksheet.min_row
        max_row = worksheet.max_row
        min_col = worksheet.min_column
        max_col = worksheet.max_column
        
        # Find actual last row with data
        for row in range(max_row, min_row - 1, -1):
            if any(worksheet.cell(row, col).value is not None for col in range(min_col, max_col + 1)):
                max_row = row
                break
        
        # Find actual last column with data
        for col in range(max_col, min_col - 1, -1):
            if any(worksheet.cell(row, col).value is not None for row in range(min_row, max_row + 1)):
                max_col = col
                break
        
        return min_row, max_row, min_col, max_col
    
    def _is_table_region(self, worksheet, start_row, start_col, end_row, end_col):
        """
        Heuristic to determine if a region contains a structured table.
        """
        # Check if first row looks like headers (all non-empty)
        first_row_filled = sum(1 for col in range(start_col, end_col + 1) 
                               if worksheet.cell(start_row, col).value is not None)
        total_cols = end_col - start_col + 1
        
        # If first row is mostly filled (>50%), likely a header row
        if first_row_filled / total_cols > 0.5:
            return True
        
        return False
    
    def _extract_table_to_markdown(self, worksheet, start_row, start_col, end_row, end_col, file_path_spaces):
        """
        Extract a table region and convert to markdown.
        Filters out empty rows/columns and 'Unnamed' columns.
        """
        self.logger.debug(f"{file_path_spaces} - Extracting table from cells "
                         f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}")
        
        md = ""
        rows_data = []
        
        # Read all rows
        for row_idx in range(start_row, end_row + 1):
            row_data = []
            for col_idx in range(start_col, end_col + 1):
                cell = worksheet.cell(row_idx, col_idx)
                # Get calculated value for formulas, not the formula itself
                value = cell.value if cell.value is not None else ""
                # Clean up the value
                value = str(value).strip().replace("\n", " ").replace("|", "\\|")  # Escape pipes
                row_data.append(value)
            
            # Skip completely empty rows
            if any(val for val in row_data):
                rows_data.append(row_data)
        
        if not rows_data:
            return ""
        
        # Check if first row is headers and filter out 'Unnamed' columns
        header_row = rows_data[0]
        valid_col_indices = []
        filtered_headers = []
        
        for idx, header in enumerate(header_row):
            # Keep column if it's not empty and doesn't start with 'Unnamed'
            if header and not header.startswith('Unnamed'):
                valid_col_indices.append(idx)
                filtered_headers.append(header)
            # Also keep columns that have data in rows below even if header is empty/unnamed
            elif any(rows_data[row_idx][idx] for row_idx in range(1, len(rows_data)) if idx < len(rows_data[row_idx])):
                valid_col_indices.append(idx)
                filtered_headers.append(header if header else f"Column {idx+1}")
        
        # If no valid columns, return empty
        if not valid_col_indices:
            return ""
        
        # Filter all rows to only include valid columns
        filtered_rows = []
        for row in rows_data:
            filtered_row = [row[idx] if idx < len(row) else "" for idx in valid_col_indices]
            # Only include rows that have at least one non-empty value
            if any(val for val in filtered_row):
                filtered_rows.append(filtered_row)
        
        # Convert to markdown table
        if filtered_rows:
            # First row as header
            md += "| " + " | ".join(filtered_rows[0]) + " |\n"
            # Separator
            md += "| " + " | ".join(["---"] * len(filtered_rows[0])) + " |\n"
            # Data rows (skip header)
            for row in filtered_rows[1:]:
                md += "| " + " | ".join(row) + " |\n"
            md += "\n"
        
        return md
    
    def _convert_chart_to_markdown(self, chart, chart_count, worksheet, sheet_name, images_folder, file_path, file_path_spaces):
        """
        Convert Excel chart to markdown representation.
        Extracts data directly from chart series, similar to PowerPoint's approach.
        """
        self.logger.debug(f"{file_path_spaces} - Processing chart number {chart_count}")
        
        md_content = f"**Chart {chart_count}**\n"
        
        # Determine chart type from class name
        chart_type = type(chart).__name__
        md_content += f"_Chart Type: {chart_type}_\n\n"
        
        try:
            # Extract chart title if available
            if hasattr(chart, 'title') and chart.title:
                try:
                    title_text = chart.title.value if hasattr(chart.title, 'value') else str(chart.title)
                    if title_text and title_text.strip():
                        md_content += f"**{title_text}**\n\n"
                except:
                    pass
            
            # Check if chart has any series
            if not chart.series or len(chart.series) == 0:
                self.logger.debug(f"{file_path_spaces} - Chart {chart_count} has no series data")
                md_content += "_Chart has no data series_\n\n"
                return md_content
            
            markdown_chart_table = []
            
            # Handle Pie/Doughnut Charts (single series, categories vs values)
            if isinstance(chart, (PieChart, PieChart3D, DoughnutChart)):
                series = chart.series[0]
                
                # Get categories - try multiple methods
                categories = self._extract_chart_categories(series, worksheet)
                # Get values
                values = self._extract_chart_values(series, worksheet)
                
                if not values:
                    self.logger.debug(f"{file_path_spaces} - No values extracted for pie/doughnut chart")
                    md_content += "_Unable to extract chart data_\n\n"
                    return md_content
                
                # Ensure we have labels
                if not categories:
                    categories = [f"Slice {i+1}" for i in range(len(values))]
                
                # Build table
                markdown_chart_table.append("| Slice Label | Value |")
                markdown_chart_table.append("|---|---|")
                for label, value in zip(categories, values):
                    markdown_chart_table.append(f"| {label} | {value} |")
            
            # Handle Scatter/Bubble Charts (x/y coordinates)
            elif isinstance(chart, (ScatterChart, BubbleChart)):
                header = "| Series | X Value | Y Value |"
                separator = "|---|---|---|"
                
                if isinstance(chart, BubbleChart):
                    header += " Bubble Size |"
                    separator += "---|"
                
                markdown_chart_table.append(header)
                markdown_chart_table.append(separator)
                
                for series in chart.series:
                    series_name = self._get_series_name(series)
                    x_values = self._extract_chart_x_values(series, worksheet)
                    y_values = self._extract_chart_y_values(series, worksheet)
                    
                    if not x_values or not y_values:
                        continue
                    
                    if isinstance(chart, BubbleChart):
                        bubble_sizes = self._extract_chart_bubble_sizes(series, worksheet)
                        for x, y, size in zip(x_values, y_values, bubble_sizes if bubble_sizes else [1]*len(x_values)):
                            markdown_chart_table.append(f"| {series_name} | {x} | {y} | {size} |")
                    else:
                        for x, y in zip(x_values, y_values):
                            markdown_chart_table.append(f"| {series_name} | {x} | {y} |")
            
            # Handle Category-based Charts (Bar, Column, Line, Area, Radar, Surface)
            elif isinstance(chart, (BarChart, BarChart3D, LineChart, LineChart3D, 
                                   AreaChart, AreaChart3D, RadarChart, 
                                   SurfaceChart, SurfaceChart3D, StockChart)):
                # Get categories from first series
                if not chart.series:
                    md_content += "_Chart has no series_\n\n"
                    return md_content
                
                first_series = chart.series[0]
                categories = self._extract_chart_categories(first_series, worksheet)
                
                # Build header row with series names
                series_names = [self._get_series_name(s) for s in chart.series]
                header_row = ["Category"] + series_names
                markdown_chart_table.append("| " + " | ".join(header_row) + " |")
                markdown_chart_table.append("| " + " | ".join(["---"] * len(header_row)) + " |")
                
                # Extract values for all series
                all_series_values = []
                for series in chart.series:
                    values = self._extract_chart_values(series, worksheet)
                    all_series_values.append(values)
                
                # Determine row count
                max_len = len(categories) if categories else max([len(v) for v in all_series_values if v], default=0)
                
                if max_len == 0:
                    md_content += "_No data could be extracted from chart_\n\n"
                    return md_content
                
                # Ensure categories list is correct length
                if not categories:
                    categories = [f"Row {i+1}" for i in range(max_len)]
                elif len(categories) < max_len:
                    categories.extend([f"Row {i+1}" for i in range(len(categories), max_len)])
                
                # Build data rows
                for idx in range(max_len):
                    row = [str(categories[idx]) if idx < len(categories) else ""]
                    for series_values in all_series_values:
                        if series_values and idx < len(series_values):
                            row.append(str(series_values[idx]))
                        else:
                            row.append("")
                    markdown_chart_table.append("| " + " | ".join(row) + " |")
            
            else:
                markdown_chart_table.append(f"_Unsupported chart type: {chart_type}_")
                self.logger.debug(f"{file_path_spaces} - Unsupported chart type: {chart_type}")
            
            # Add chart markdown to content
            if markdown_chart_table:
                md_content += "\n".join(markdown_chart_table) + "\n\n"
            else:
                md_content += "_No chart data available_\n\n"
        
        except Exception as e:
            self.logger.error(f"{file_path_spaces} - Error processing chart {chart_count}: {e}")
            import traceback
            self.logger.debug(f"Traceback: {traceback.format_exc()}")
            md_content += f"_Error extracting chart data_\n\n"
        
        return md_content
    
    def _get_series_name(self, series):
        """Extract series name from chart series object."""
        try:
            if hasattr(series, 'title') and series.title:
                if hasattr(series.title, 'value') and series.title.value:
                    return str(series.title.value)
                elif hasattr(series.title, 'v') and series.title.v:
                    return str(series.title.v)
            return "Series"
        except:
            return "Series"
    
    def _extract_chart_categories(self, series, worksheet):
        """Extract category labels from chart series."""
        try:
            if not hasattr(series, 'cat') or not series.cat:
                return []
            
            # Try to get from cache first
            if hasattr(series.cat, 'strRef') and series.cat.strRef:
                if hasattr(series.cat.strRef, 'strCache') and series.cat.strRef.strCache:
                    if hasattr(series.cat.strRef.strCache, 'pt') and series.cat.strRef.strCache.pt:
                        return [pt.v for pt in series.cat.strRef.strCache.pt if pt and hasattr(pt, 'v')]
            
            if hasattr(series.cat, 'numRef') and series.cat.numRef:
                if hasattr(series.cat.numRef, 'numCache') and series.cat.numRef.numCache:
                    if hasattr(series.cat.numRef.numCache, 'pt') and series.cat.numRef.numCache.pt:
                        return [pt.v for pt in series.cat.numRef.numCache.pt if pt and hasattr(pt, 'v')]
            
            # Try to read from formula reference
            formula = None
            if hasattr(series.cat, 'strRef') and series.cat.strRef and hasattr(series.cat.strRef, 'f'):
                formula = series.cat.strRef.f
            elif hasattr(series.cat, 'numRef') and series.cat.numRef and hasattr(series.cat.numRef, 'f'):
                formula = series.cat.numRef.f
            
            if formula:
                return self._parse_formula_reference(formula, worksheet)
            
        except Exception as e:
            self.logger.debug(f"Error extracting categories: {e}")
        
        return []
    
    def _extract_chart_values(self, series, worksheet):
        """Extract data values from chart series."""
        try:
            if not hasattr(series, 'val') or not series.val:
                return []
            
            # Try to get from cache first
            if hasattr(series.val, 'numRef') and series.val.numRef:
                if hasattr(series.val.numRef, 'numCache') and series.val.numRef.numCache:
                    if hasattr(series.val.numRef.numCache, 'pt') and series.val.numRef.numCache.pt:
                        return [pt.v for pt in series.val.numRef.numCache.pt if pt and hasattr(pt, 'v')]
            
            # Try to read from formula reference
            if hasattr(series.val, 'numRef') and series.val.numRef and hasattr(series.val.numRef, 'f'):
                formula = series.val.numRef.f
                if formula:
                    return self._parse_formula_reference(formula, worksheet)
            
        except Exception as e:
            self.logger.debug(f"Error extracting values: {e}")
        
        return []
    
    def _extract_chart_x_values(self, series, worksheet):
        """Extract X values for scatter/bubble charts."""
        try:
            if not hasattr(series, 'xVal') or not series.xVal:
                return []
            
            # Try cache
            if hasattr(series.xVal, 'numRef') and series.xVal.numRef:
                if hasattr(series.xVal.numRef, 'numCache') and series.xVal.numRef.numCache:
                    if hasattr(series.xVal.numRef.numCache, 'pt') and series.xVal.numRef.numCache.pt:
                        return [pt.v for pt in series.xVal.numRef.numCache.pt if pt and hasattr(pt, 'v')]
            
            # Try formula
            if hasattr(series.xVal, 'numRef') and series.xVal.numRef and hasattr(series.xVal.numRef, 'f'):
                return self._parse_formula_reference(series.xVal.numRef.f, worksheet)
        
        except Exception as e:
            self.logger.debug(f"Error extracting X values: {e}")
        
        return []
    
    def _extract_chart_y_values(self, series, worksheet):
        """Extract Y values for scatter/bubble charts."""
        try:
            if not hasattr(series, 'yVal') or not series.yVal:
                return []
            
            # Try cache
            if hasattr(series.yVal, 'numRef') and series.yVal.numRef:
                if hasattr(series.yVal.numRef, 'numCache') and series.yVal.numRef.numCache:
                    if hasattr(series.yVal.numRef.numCache, 'pt') and series.yVal.numRef.numCache.pt:
                        return [pt.v for pt in series.yVal.numRef.numCache.pt if pt and hasattr(pt, 'v')]
            
            # Try formula
            if hasattr(series.yVal, 'numRef') and series.yVal.numRef and hasattr(series.yVal.numRef, 'f'):
                return self._parse_formula_reference(series.yVal.numRef.f, worksheet)
        
        except Exception as e:
            self.logger.debug(f"Error extracting Y values: {e}")
        
        return []
    
    def _extract_chart_bubble_sizes(self, series, worksheet):
        """Extract bubble sizes for bubble charts."""
        try:
            if not hasattr(series, 'bubbleSize') or not series.bubbleSize:
                return []
            
            # Try cache
            if hasattr(series.bubbleSize, 'numRef') and series.bubbleSize.numRef:
                if hasattr(series.bubbleSize.numRef, 'numCache') and series.bubbleSize.numRef.numCache:
                    if hasattr(series.bubbleSize.numRef.numCache, 'pt') and series.bubbleSize.numRef.numCache.pt:
                        return [pt.v for pt in series.bubbleSize.numRef.numCache.pt if pt and hasattr(pt, 'v')]
            
            # Try formula
            if hasattr(series.bubbleSize, 'numRef') and series.bubbleSize.numRef and hasattr(series.bubbleSize.numRef, 'f'):
                return self._parse_formula_reference(series.bubbleSize.numRef.f, worksheet)
        
        except Exception as e:
            self.logger.debug(f"Error extracting bubble sizes: {e}")
        
        return []
    
    def _parse_formula_reference(self, formula, worksheet):
        """
        Parse a formula reference string like 'Sheet1!$A$1:$A$10' and extract values.
        """
        values = []
        try:
            if not formula:
                return values
            
            # Remove sheet name if present
            if '!' in formula:
                formula = formula.split('!')[-1]
            
            # Remove $ signs
            formula = formula.replace('$', '')
            
            # Parse range
            if ':' in formula:
                from openpyxl.utils import coordinate_to_tuple
                
                parts = formula.split(':')
                if len(parts) != 2:
                    return values
                
                start, end = parts
                start_row, start_col = coordinate_to_tuple(start)
                end_row, end_col = coordinate_to_tuple(end)
                
                # Read values from cells
                for row in range(start_row, end_row + 1):
                    for col in range(start_col, end_col + 1):
                        cell_value = worksheet.cell(row, col).value
                        if cell_value is not None:
                            values.append(cell_value)
            else:
                # Single cell reference
                from openpyxl.utils import coordinate_to_tuple
                row, col = coordinate_to_tuple(formula)
                cell_value = worksheet.cell(row, col).value
                if cell_value is not None:
                    values.append(cell_value)
                    
        except Exception as e:
            self.logger.debug(f"Error parsing formula reference '{formula}': {e}")
        
        return values
    
    def _extract_images_to_markdown(self, worksheet, images_folder, file_path, sheet_name, file_path_spaces):
        """
        Extract images from worksheet and convert to markdown.
        """
        md_content = ""
        
        if hasattr(worksheet, '_images') and worksheet._images:
            for img_idx, img in enumerate(worksheet._images):
                try:
                    figure_count = img_idx + 1
                    self.logger.debug(f"{file_path_spaces} - Processing image {figure_count} in sheet '{sheet_name}'")
                    
                    # Get image data
                    img_data = img._data()
                    image = Image.open(BytesIO(img_data))
                    
                    # Check image size
                    img_width_px, img_height_px = image.size
                    image_area_px = img_width_px * img_height_px
                    
                    min_image_size = 500
                    if image_area_px < min_image_size:
                        self.logger.debug(f"{file_path_spaces} - Skipping small image {figure_count} "
                                        f"({img_width_px}x{img_height_px} = {image_area_px}, threshold: {min_image_size})")
                        continue
                    
                    # Save image
                    img_path = Path(f"{images_folder}/{file_path.name.replace(' ', '-')}-{sheet_name.replace(' ', '-')}-{figure_count}.png")
                    image.save(img_path)
                    self.logger.debug(f"{file_path_spaces} - Image saved at {img_path}")
                    
                    # Add to markdown
                    md_content += f"**Figure {figure_count}**\n"
                    self.logger.debug(f"{file_path_spaces} - Replacing image with markdown using GPT vision")
                    md_content += self._gpt4o_vision(img_path)
                    md_content += "\n\n"
                
                except Exception as e:
                    self.logger.error(f"{file_path_spaces} - Error processing image {img_idx + 1}: {e}")
        
        return md_content
    
    def _capture_sheet_with_charts(self, worksheet, sheet_name, images_folder, file_path, file_path_spaces):
        """
        Alternative method: If charts exist but can't be extracted programmatically,
        capture them as images and use vision to interpret them.
        This requires additional libraries like xlwings or win32com (Windows only) or 
        converting to PDF first. For now, this is a placeholder for the approach.
        """
        # NOTE: This would require external tools to render Excel to image
        # Options include:
        # 1. xlwings (Windows/Mac with Excel installed)
        # 2. unoconv + LibreOffice
        # 3. Excel to PDF then PDF to images
        # For MVP, we'll rely on programmatic extraction
        pass
    
    def _process_sheet(self, worksheet, sheet_name, images_folder, file_path, file_path_spaces):
        """
        Process a single worksheet and convert to markdown.
        """
        self.logger.info(f"{file_path_spaces} - Processing sheet: '{sheet_name}'")
        
        md_content = f"# {sheet_name}\n\n"
        
        # Get used range
        used_range = self._get_used_range(worksheet)
        if not used_range:
            self.logger.info(f"{file_path_spaces} - Sheet '{sheet_name}' is empty, skipping")
            return ""
        
        min_row, max_row, min_col, max_col = used_range
        
        # Extract main table/data
        if self._is_table_region(worksheet, min_row, min_col, max_row, max_col):
            md_content += self._extract_table_to_markdown(worksheet, min_row, min_col, max_row, max_col, file_path_spaces)
        else:
            # If not a clear table, still output as table format
            md_content += self._extract_table_to_markdown(worksheet, min_row, min_col, max_row, max_col, file_path_spaces)
        
        # Process charts - check multiple sources
        chart_count = 0
        
        # Method 1: worksheet._charts (standard openpyxl)
        if hasattr(worksheet, '_charts') and worksheet._charts:
            self.logger.info(f"{file_path_spaces} - Found {len(worksheet._charts)} chart(s) via _charts in sheet '{sheet_name}'")
            for chart_idx, chart in enumerate(worksheet._charts):
                try:
                    chart_md = self._convert_chart_to_markdown(
                        chart, chart_count + 1, worksheet, sheet_name, 
                        images_folder, file_path, file_path_spaces
                    )
                    if chart_md:
                        md_content += chart_md
                        chart_count += 1
                except Exception as e:
                    self.logger.error(f"{file_path_spaces} - Error processing chart {chart_idx + 1} in sheet '{sheet_name}': {e}")
        
        # Method 2: Check worksheet drawings for embedded charts
        if hasattr(worksheet, '_drawing') and worksheet._drawing:
            self.logger.debug(f"{file_path_spaces} - Checking drawings for charts in sheet '{sheet_name}'")
            # Drawing objects can contain charts that aren't in _charts
            # This is a workaround for charts that openpyxl doesn't fully parse
        
        if chart_count == 0:
            self.logger.debug(f"{file_path_spaces} - No charts found via standard methods in sheet '{sheet_name}'")
        
        # Process images
        md_content += self._extract_images_to_markdown(worksheet, images_folder, file_path, sheet_name, file_path_spaces)
        
        return md_content
    
    def _extract_macro_info(self, file_path, file_path_spaces):
        """
        Extract macro information from xlsm files.
        Note: openpyxl cannot execute or fully read VBA code, but we can detect its presence.
        """
        md_content = ""
        
        try:
            # Check if file is macro-enabled
            if file_path.suffix.lower() in ['.xlsm', '.xlam']:
                wb = openpyxl.load_workbook(file_path, keep_vba=True)
                
                if wb.vba_archive:
                    self.logger.info(f"{file_path_spaces} - Macro-enabled workbook detected")
                    md_content += "## Macro Information\n\n"
                    md_content += "_This workbook contains VBA macros. "
                    md_content += "Macro code cannot be fully extracted by this parser, but its presence is noted._\n\n"
                    
                    # Try to list macro modules if possible
                    try:
                        # This is limited - openpyxl doesn't fully expose VBA structure
                        md_content += "_Macro modules detected in workbook_\n\n"
                    except:
                        pass
        
        except Exception as e:
            self.logger.debug(f"{file_path_spaces} - Could not extract macro info: {e}")
        
        return md_content
    
    def parse(self, file_path: Path, file_path_spaces: str) -> List[dict]:
        """
        Parse Excel file and convert to markdown documents per sheet.
        """
        documents = []
        
        with tempfile.TemporaryDirectory(prefix="excel_img_render_") as tmp_img_dir:
            images_folder = Path(tmp_img_dir)
            
            try:
                self.logger.info(f"{file_path_spaces} - Extracting content from Excel file... ")
                
                # Load workbook with data_only=True to get calculated values instead of formulas
                wb = openpyxl.load_workbook(file_path, data_only=True, keep_vba=True)
                
                # Extract macro info if present
                macro_info = self._extract_macro_info(file_path, file_path_spaces)
                
                total_sheets = len(wb.sheetnames)
                
                for sheet_idx, sheet_name in enumerate(wb.sheetnames):
                    self.logger.info(f"{file_path_spaces} - **Sheet {sheet_idx + 1}/{total_sheets}: '{sheet_name}'**...")
                    
                    try:
                        worksheet = wb[sheet_name]
                        
                        # Process sheet
                        sheet_markdown = self._process_sheet(worksheet, sheet_name, images_folder, file_path, file_path_spaces)
                        
                        # Add macro info to first sheet if present
                        if sheet_idx == 0 and macro_info:
                            sheet_markdown = macro_info + sheet_markdown
                        
                        # Skip empty sheets
                        if not sheet_markdown.strip() or sheet_markdown.strip() == f"# {sheet_name}":
                            self.logger.info(f"{file_path_spaces} - Sheet '{sheet_name}' has no content, skipping")
                            continue
                        
                        metadata = {
                            "file_path": file_path_spaces,
                            "sheet": sheet_name,
                            "sheet_number": sheet_idx + 1,
                            "file_hash": compute_file_hash(file_path)
                        }
                        metadata.update(self._load_sidecar_metadata(file_path))
                        
                        documents.append({"markdown": sheet_markdown, "metadata": metadata})
                    
                    except Exception as e:
                        self.logger.error(f"{file_path_spaces} - X Error on sheet '{sheet_name}': {e}")
                        continue
                
                self.logger.success(f"✓ Converted {file_path_spaces}")
                return documents
            
            except Exception as e:
                self.logger.error(f"Failed to process {file_path_spaces}: {e}")
                return []
