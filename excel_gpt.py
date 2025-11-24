from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

# Chartsheet support – backward compatible
try:
    from openpyxl.worksheet.chartsheet import ChartSheet
except ImportError:
    ChartSheet = None

from PIL import Image
from io import BytesIO


class ExcelParser(BaseParser):
    def __init__(self) -> None:
        super().__init__(name="EXCEL_PARSER")

    # -------------------------------------------------------------
    # Core parse() method
    # -------------------------------------------------------------
    def parse(self, file_path: Path, file_path_spaces: str) -> List[dict]:
        try:
            # Use values ONLY (NO formulas)
            workbook = load_workbook(filename=file_path, data_only=True)
        except Exception as exc:
            self.logger.error(f"✗ Failed to open {file_path_spaces}: {exc}")
            return []

        documents: List[dict] = []
        file_hash = compute_file_hash(file_path)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            metadata_base = {
                "file_path": file_path,
                "file_path_spaces": file_path_spaces,
                "sheet_name": sheet_name,
                "file_hash": file_hash,
            }
            metadata_base.update(self._load_sidecar_metadata(file_path))

            # 🟦 1. CHARTSHEET – only extract charts
            if ChartSheet and isinstance(sheet, ChartSheet):
                documents.extend(self._build_chart_documents(sheet, metadata_base))
                continue

            # 🟩 2. NORMAL WORKSHEET
            if isinstance(sheet, Worksheet):
                df = self._sheet_to_dataframe(sheet)
                if df is not None:
                    documents.extend(self._build_table_documents(df, metadata_base))

                documents.extend(self._build_excel_table_documents(sheet, metadata_base))
                documents.extend(self._build_chart_documents(sheet, metadata_base))
                documents.extend(self._build_image_documents(sheet, metadata_base))

        self.logger.info(f"✓ Parsed {file_path_spaces} with {len(documents)} fragments")
        return documents

    # -------------------------------------------------------------
    #  Excel → DataFrame conversion
    # -------------------------------------------------------------
    def _sheet_to_dataframe(self, sheet) -> Optional[pd.DataFrame]:
        if sheet.max_row == 0 or sheet.max_column == 0:
            return None

        # Build matrix including merged cells
        matrix = [[None for _ in range(sheet.max_column)] for _ in range(sheet.max_row)]
        for row in sheet.iter_rows():
            for cell in row:
                matrix[cell.row - 1][cell.column - 1] = cell.value

        # Fill merged cells with top-left value (preserves table structure)
        for merged_range in sheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = (
                merged_range.min_col,
                merged_range.min_row,
                merged_range.max_col,
                merged_range.max_row,
            )
            value = sheet.cell(min_row, min_col).value
            for r in range(min_row - 1, max_row):
                for c in range(min_col - 1, max_col):
                    if matrix[r][c] is None:
                        matrix[r][c] = value

        df = pd.DataFrame(matrix)
        df = df.dropna(how="all").dropna(axis=1, how="all")

        if df.empty:
            return None

        # Promote first row to header if valid text exists
        header = df.iloc[0]
        if header.notna().any():
            df.columns = header
            df = df[1:]

        df = df.reset_index(drop=True)
        return df

    # -------------------------------------------------------------
    # 🧹 Clean DataFrame before markdown conversion
    # -------------------------------------------------------------
    def _df_to_markdown_clean(self, df: pd.DataFrame) -> str:
        cleaned = df.copy()
        cleaned = cleaned.dropna(how="all").dropna(axis=1, how="all")
        cleaned = cleaned.loc[:, ~cleaned.columns.astype(str).str.match(r"^Unnamed")]
        cleaned = cleaned.where(pd.notnull(cleaned), "")
        cleaned.columns = [("" if c is None else str(c)) for c in cleaned.columns]
        return cleaned.to_markdown(index=False)

    # -------------------------------------------------------------
    #  Build table-level documents
    # -------------------------------------------------------------
    def _build_table_documents(
        self, df: pd.DataFrame, metadata_base: Dict[str, Any]
    ) -> List[dict]:
        docs = []
        md_full = self._df_to_markdown_clean(df)

        docs.append({
            "markdown": md_full,
            "metadata": {**metadata_base, "type": "full_sheet"}
        })

        # chunk the table for better embedding
        for idx, chunk_df in enumerate(self._chunk_dataframe(df, max_rows=40)):
            docs.append({
                "markdown": self._df_to_markdown_clean(chunk_df),
                "metadata": {**metadata_base, "type": "chunk", "chunk_index": idx}
            })

        return docs

    def _chunk_dataframe(self, df: pd.DataFrame, max_rows: int) -> List[pd.DataFrame]:
        return [df.iloc[i:i + max_rows] for i in range(0, len(df), max_rows)]

    # -------------------------------------------------------------
    # 📋 Formal Excel Tables (with caption)
    # -------------------------------------------------------------
    def _find_caption_for_object(self, sheet, row_index: int) -> Optional[str]:
        """Find text nearest to the given row in worksheet."""
        min_distance = float("inf")
        caption = None

        for r, row in enumerate(sheet.iter_rows(values_only=True)):
            for cell in row:
                if isinstance(cell, str) and cell.strip():
                    dist = abs(r - row_index)
                    if dist < min_distance:
                        min_distance = dist
                        caption = cell.strip()
        return caption

    def _build_excel_table_documents(self, sheet, metadata_base) -> List[dict]:
        docs = []
        for table in getattr(sheet, "_tables", []):
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            rows = [
                [cell.value for cell in row]
                for row in sheet.iter_rows(min_row=min_row, max_row=max_row,
                                           min_col=min_col, max_col=max_col)
            ]

            df = pd.DataFrame(rows)
            df.columns = df.iloc[0]
            df = df[1:].reset_index(drop=True)

            caption = self._find_caption_for_object(sheet, min_row - 2)
            md = f"**Table: {caption}**\n" if caption else "**Table**\n"
            md += self._df_to_markdown_clean(df)

            docs.append({
                "markdown": md,
                "metadata": {**metadata_base, "type": "excel_table", "table_name": table.name}
            })

        return docs

    # -------------------------------------------------------------
    # 📈 Chart documents with captions
    # -------------------------------------------------------------
    def _build_chart_documents(self, sheet, metadata_base) -> List[dict]:
        docs = []
        charts = getattr(sheet, "_charts", [])
        for i, chart in enumerate(charts, start=1):
            try:
                row_index = chart.anchor._from.row
            except Exception:
                row_index = 0

            caption = self._find_caption_for_object(sheet, row_index)
            md = f"**Chart {i}: {caption}**\n" if caption else f"**Chart {i}**\n"
            md += f"_Chart Type: {chart.__class__.__name__}_\n"

            docs.append({
                "markdown": md,
                "metadata": {**metadata_base, "type": "embedded_chart", "index": i}
            })

        return docs

    # -------------------------------------------------------------
    # 🖼 Extract embedded images with caption (Ppt logic)
    # -------------------------------------------------------------
    def _build_image_documents(self, sheet, metadata_base) -> List[dict]:
        docs = []
        images = getattr(sheet, "_images", {})
        for i, img in enumerate(images, start=1):
            try:
                row_index = img.anchor._from.row
            except:
                row_index = 0

            caption = self._find_caption_for_object(sheet, row_index)
            md = f"**Figure {i}: {caption}**\n" if caption else f"**Figure {i}**\n"

            docs.append({
                "markdown": md,
                "metadata": {**metadata_base, "type": "embedded_image", "index": i}
            })

        return docs
