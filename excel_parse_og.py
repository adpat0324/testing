import hashlib
import logging
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries


def compute_file_hash(file_path: Path) -> str:
    """Compute a stable hash for a file on disk."""
    sha = hashlib.sha256()
    with file_path.open("rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            sha.update(chunk)
    return sha.hexdigest()


class BaseParser:  # Minimal fallback; replace with shared implementation if available
    def __init__(self, name: str) -> None:
        self.name = name
        self.logger = logging.getLogger(name)

    def _load_sidecar_metadata(self, _: Path) -> Dict[str, Any]:
        return {}


class ExcelParser(BaseParser):
    def __init__(self) -> None:
        super().__init__(name="EXCEL_PARSER")

    def parse(self, file_path: Path, file_path_spaces: str) -> List[dict]:
        try:
            workbook = load_workbook(filename=file_path, data_only=True)
        except Exception as exc:  # pragma: no cover - logging path
            self.logger.error(f"✗ Failed to open {file_path_spaces}: {exc}")
            return []

        documents: List[dict] = []
        file_hash = compute_file_hash(file_path)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            dataframe = self._sheet_to_dataframe(sheet)

            metadata_base = {
                "file_path": file_path,
                "file_path_spaces": file_path_spaces,
                "sheet_name": sheet_name,
                "file_hash": file_hash,
            }
            metadata_base.update(self._load_sidecar_metadata(file_path))

            if dataframe is not None and not dataframe.empty:
                documents.extend(self._build_table_documents(dataframe, metadata_base))

            documents.extend(self._build_excel_table_documents(sheet, metadata_base))
            documents.extend(self._build_image_documents(sheet, metadata_base))
            documents.extend(self._build_chart_documents(sheet, metadata_base))

        self.logger.info(f"✓ Parsed {file_path_spaces} with {len(documents)} fragments")
        return documents

    # ------------------------------------------------------------------
    # Sheet helpers
    # ------------------------------------------------------------------
    @staticmethod
    def _sheet_to_dataframe(sheet) -> Optional[pd.DataFrame]:
        """Convert a worksheet to a cleaned pandas DataFrame.

        * Preserves merged-cell values by forward filling the merged range.
        * Drops fully empty rows/columns.
        * Treats the first non-empty row as a header if present.
        """

        if sheet.max_row == 0 or sheet.max_column == 0:
            return None

        matrix: List[List[Any]] = [[None for _ in range(sheet.max_column)] for _ in range(sheet.max_row)]
        for row in sheet.iter_rows():
            for cell in row:
                matrix[cell.row - 1][cell.column - 1] = cell.value

        # Fill merged cell ranges with their top-left value
        for merged_range in sheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = (
                merged_range.min_col,
                merged_range.min_row,
                merged_range.max_col,
                merged_range.max_row,
            )
            value = sheet.cell(min_row, min_col).value
            for r in range(min_row - 1, max_row):
                for c in range(min_col - 1, max_col):
                    if matrix[r][c] is None:
                        matrix[r][c] = value

        df = pd.DataFrame(matrix)
        df = df.dropna(how="all").dropna(axis=1, how="all")
        if df.empty:
            return None

        # Promote first row to header if it contains any strings
        header_candidates = df.iloc[0]
        if header_candidates.notna().any() and header_candidates.astype(str).str.len().any():
            df.columns = header_candidates
            df = df[1:]

        df = df.reset_index(drop=True)
        return df

    def _build_table_documents(self, df: pd.DataFrame, metadata_base: Dict[str, Any]) -> List[dict]:
        documents: List[dict] = []

        markdown_full = df.to_markdown(index=False)
        metadata_full = {**metadata_base, "type": "full_sheet", "num_rows": len(df), "num_columns": len(df.columns)}
        documents.append({
            "markdown": markdown_full,
            "metadata": metadata_full,
        })

        for idx, chunk in enumerate(self._chunk_dataframe(df, max_rows=40)):
            documents.append({
                "markdown": chunk,
                "metadata": {
                    **metadata_base,
                    "type": "chunk",
                    "chunk_index": idx,
                    "num_rows": len(df),
                    "num_columns": len(df.columns),
                },
            })

        return documents

    @staticmethod
    def _chunk_dataframe(df: pd.DataFrame, max_rows: int = 30) -> List[str]:
        chunks: List[str] = []
        for start in range(0, len(df), max_rows):
            chunk_df = df.iloc[start:start + max_rows]
            chunks.append(chunk_df.to_markdown(index=False))
        return chunks

    # ------------------------------------------------------------------
    # Embedded content helpers
    # ------------------------------------------------------------------
    def _build_excel_table_documents(self, sheet, metadata_base: Dict[str, Any]) -> List[dict]:
        documents: List[dict] = []
        for table in getattr(sheet, "_tables", []):
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            values = []
            for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                values.append([cell.value for cell in row])

            df = pd.DataFrame(values)
            if df.empty:
                continue
            # Use first row as header for formal tables
            df.columns = df.iloc[0]
            df = df[1:].reset_index(drop=True)

            documents.append({
                "markdown": df.to_markdown(index=False),
                "metadata": {
                    **metadata_base,
                    "type": "excel_table",
                    "table_name": table.name,
                    "table_range": table.ref,
                    "num_rows": len(df),
                    "num_columns": len(df.columns),
                },
            })
        return documents

    def _build_image_documents(self, sheet, metadata_base: Dict[str, Any]) -> List[dict]:
        documents: List[dict] = []
        images = getattr(sheet, "_images", [])
        if not images:
            return documents

        image_rows = []
        for idx, image in enumerate(images, start=1):
            anchor = getattr(image, "anchor", None)
            anchor_text = getattr(anchor, "_from", None) or anchor
            image_rows.append({
                "#": idx,
                "location": str(anchor_text),
                "format": getattr(image, "format", "unknown"),
            })

        df = pd.DataFrame(image_rows)
        documents.append({
            "markdown": df.to_markdown(index=False),
            "metadata": {**metadata_base, "type": "embedded_images", "count": len(images)},
        })
        return documents

    def _build_chart_documents(self, sheet, metadata_base: Dict[str, Any]) -> List[dict]:
        documents: List[dict] = []
        charts = getattr(sheet, "_charts", [])
        if not charts:
            return documents

        rows = []
        for idx, chart in enumerate(charts, start=1):
            title = None
            if hasattr(chart, "title"):
                try:
                    title = chart.title if not hasattr(chart.title, "tx") else chart.title.tx.rich.p[0].r[0].t
                except Exception:
                    title = getattr(chart.title, "_text", None)
            rows.append({
                "#": idx,
                "chart_type": chart.__class__.__name__,
                "title": title or "(no title)",
            })

        df = pd.DataFrame(rows)
        documents.append({
            "markdown": df.to_markdown(index=False),
            "metadata": {**metadata_base, "type": "embedded_charts", "count": len(charts)},
        })
        return documents
