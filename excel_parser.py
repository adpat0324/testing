from pathlib import Path
from typing import List, Dict, Any, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries


class ExcelParser(BaseParser):
    def __init__(self) -> None:
        super().__init__(name="EXCEL_PARSER")

    def parse(self, file_path: Path, file_path_spaces: str) -> List[dict]:
        """Parse an Excel file into RAG-friendly markdown fragments."""
        try:
            # One copy with computed values, one with formulas
            wb_values = load_workbook(filename=file_path, data_only=True)
            wb_formulas = load_workbook(filename=file_path, data_only=False)
        except Exception as exc:  # pragma: no cover - logging path
            self.logger.error(f"✗ Failed to open {file_path_spaces}: {exc}")
            return []

        documents: List[dict] = []
        file_hash = compute_file_hash(file_path)

        for sheet_name in wb_values.sheetnames:
            sheet_vals = wb_values[sheet_name]
            sheet_formulas = wb_formulas[sheet_name]

            metadata_base = {
                "file_path": file_path,
                "file_path_spaces": file_path_spaces,
                "sheet_name": sheet_name,
                "file_hash": file_hash,
            }
            metadata_base.update(self._load_sidecar_metadata(file_path))

            # --------- main sheet as dataframe / table ----------
            df = self._sheet_to_dataframe(sheet_vals)
            if df is not None and not df.empty:
                documents.extend(self._build_table_documents(df, metadata_base))

            # formal Excel tables
            documents.extend(self._build_excel_table_documents(sheet_vals, metadata_base))
            # embedded images
            documents.extend(self._build_image_documents(sheet_vals, metadata_base))
            # embedded charts
            documents.extend(self._build_chart_documents(sheet_vals, metadata_base))
            # formulas (from formulas workbook)
            documents.extend(self._build_formula_documents(sheet_formulas, metadata_base))

        # macros live at workbook level, not per-sheet
        metadata_workbook = {
            "file_path": file_path,
            "file_path_spaces": file_path_spaces,
            "sheet_name": "(workbook)",
            "file_hash": file_hash,
        }
        metadata_workbook.update(self._load_sidecar_metadata(file_path))
        documents.extend(self._build_vba_documents(file_path, metadata_workbook))

        self.logger.info(f"✓ Parsed {file_path_spaces} with {len(documents)} fragments")
        return documents

    # ------------------------------------------------------------------
    # Sheet helpers
    # ------------------------------------------------------------------
    @staticmethod
    def _sheet_to_dataframe(sheet) -> Optional[pd.DataFrame]:
        """Convert a worksheet to a cleaned pandas DataFrame.

        * Preserves merged-cell values by forward filling the merged range.
        * Drops fully empty rows/columns.
        * Uses the first non-empty row as header if it looks like one.
        """
        if sheet.max_row == 0 or sheet.max_column == 0:
            return None

        # Build full matrix
        matrix: List[List[Any]] = [
            [None for _ in range(sheet.max_column)] for _ in range(sheet.max_row)
        ]
        for row in sheet.iter_rows():
            for cell in row:
                matrix[cell.row - 1][cell.column - 1] = cell.value

        # Fill merged cell ranges with their top-left value
        for merged_range in sheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = (
                merged_range.min_col,
                merged_range.min_row,
                merged_range.max_col,
                merged_range.max_row,
            )
            value = sheet.cell(min_row, min_col).value
            for r in range(min_row - 1, max_row):
                for c in range(min_col - 1, max_col):
                    if matrix[r][c] is None:
                        matrix[r][c] = value

        df = pd.DataFrame(matrix)

        # Drop completely empty rows/cols
        df = df.dropna(how="all").dropna(axis=1, how="all")
        if df.empty:
            return None

        # Promote first row to header if it looks like labels
        header = df.iloc[0]
        if header.notna().any() and header.astype(str).str.len().any():
            df.columns = header
            df = df[1:]

        df = df.reset_index(drop=True)
        return df

    # Small helper to avoid NaN/Unnamed noise in markdown
    @staticmethod
    def _df_to_markdown_clean(df: pd.DataFrame) -> str:
        if df is None or df.empty:
            return ""
        cleaned = df.copy()
        # Remove completely blank cols/rows again just in case
        cleaned = cleaned.dropna(how="all").dropna(axis=1, how="all")
        # Drop "Unnamed: ..." style columns
        cleaned = cleaned.loc[
            :, ~cleaned.columns.astype(str).str.match(r"^Unnamed(:\s*\d+)?$")
        ]
        # Replace NaN with empty string
        cleaned = cleaned.where(pd.notnull(cleaned), "")
        # Stringify headers
        cleaned.columns = [("" if c is None else str(c)) for c in cleaned.columns]
        return cleaned.to_markdown(index=False)

    # ------------------------------------------------------------------
    # Main sheet table documents
    # ------------------------------------------------------------------
    def _build_table_documents(
        self, df: pd.DataFrame, metadata_base: Dict[str, Any]
    ) -> List[dict]:
        documents: List[dict] = []

        # Full-sheet representation
        markdown_full = self._df_to_markdown_clean(df)
        metadata_full = {
            **metadata_base,
            "type": "full_sheet",
            "num_rows": len(df),
            "num_columns": len(df.columns),
        }
        documents.append({"markdown": markdown_full, "metadata": metadata_full})

        # Chunked rows for RAG embeddings
        for idx, chunk_df in enumerate(self._chunk_dataframe(df, max_rows=40)):
            documents.append(
                {
                    "markdown": self._df_to_markdown_clean(chunk_df),
                    "metadata": {
                        **metadata_base,
                        "type": "chunk",
                        "chunk_index": idx,
                        "num_rows": len(df),
                        "num_columns": len(df.columns),
                    },
                }
            )

        return documents

    @staticmethod
    def _chunk_dataframe(
        df: pd.DataFrame, max_rows: int = 30
    ) -> List[pd.DataFrame]:
        chunks: List[pd.DataFrame] = []
        for start in range(0, len(df), max_rows):
            chunks.append(df.iloc[start : start + max_rows])
        return chunks

    # ------------------------------------------------------------------
    # Embedded formal Excel tables
    # ------------------------------------------------------------------
    def _build_excel_table_documents(
        self, sheet, metadata_base: Dict[str, Any]
    ) -> List[dict]:
        documents: List[dict] = []
        for table in getattr(sheet, "_tables", []):
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            values = []
            for row in sheet.iter_rows(
                min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
            ):
                values.append([cell.value for cell in row])

            df = pd.DataFrame(values)
            if df.empty:
                continue

            # First row is the header for formal tables
            df.columns = df.iloc[0]
            df = df[1:].reset_index(drop=True)

            documents.append(
                {
                    "markdown": self._df_to_markdown_clean(df),
                    "metadata": {
                        **metadata_base,
                        "type": "excel_table",
                        "table_name": table.name,
                        "table_range": table.ref,
                        "num_rows": len(df),
                        "num_columns": len(df.columns),
                    },
                }
            )
        return documents

    # ------------------------------------------------------------------
    # Embedded images
    # ------------------------------------------------------------------
    def _build_image_documents(
        self, sheet, metadata_base: Dict[str, Any]
    ) -> List[dict]:
        documents: List[dict] = []
        images = getattr(sheet, "_images", [])
        if not images:
            return documents

        rows = []
        for idx, image in enumerate(images, start=1):
            anchor = getattr(image, "anchor", None)
            anchor_text = getattr(anchor, "_from", None) or anchor
            rows.append(
                {
                    "#": idx,
                    "location": str(anchor_text),
                    "format": getattr(image, "format", "unknown"),
                }
            )

        df = pd.DataFrame(rows)
        documents.append(
            {
                "markdown": self._df_to_markdown_clean(df),
                "metadata": {
                    **metadata_base,
                    "type": "embedded_images",
                    "count": len(images),
                },
            }
        )
        return documents

    # ------------------------------------------------------------------
    # Embedded charts
    # ------------------------------------------------------------------
    def _build_chart_documents(
        self, sheet, metadata_base: Dict[str, Any]
    ) -> List[dict]:
        documents: List[dict] = []
        charts = getattr(sheet, "_charts", [])
        if not charts:
            return documents

        rows = []
        for idx, chart in enumerate(charts, start=1):
            title = None
            if hasattr(chart, "title"):
                try:
                    # try a couple of title styles
                    title = getattr(chart.title, "_text", None) or chart.title
                except Exception:
                    title = None

            rows.append(
                {
                    "#": idx,
                    "chart_type": chart.__class__.__name__,
                    "title": title or "(no title)",
                }
            )

        df = pd.DataFrame(rows)
        documents.append(
            {
                "markdown": self._df_to_markdown_clean(df),
                "metadata": {
                    **metadata_base,
                    "type": "embedded_charts",
                    "count": len(charts),
                },
            }
        )
        return documents

    # ------------------------------------------------------------------
    # Formulas
    # ------------------------------------------------------------------
    def _build_formula_documents(
        self, sheet, metadata_base: Dict[str, Any]
    ) -> List[dict]:
        """Capture formulas and their locations for RAG."""
        rows = []
        for row in sheet.iter_rows():
            for cell in row:
                # openpyxl marks formula cells with data_type 'f'
                value = cell.value
                if cell.data_type == "f" or (isinstance(value, str) and value.startswith("=")):
                    rows.append(
                        {
                            "cell": cell.coordinate,
                            "formula": str(value),
                        }
                    )

        if not rows:
            return []

        df = pd.DataFrame(rows)
        return [
            {
                "markdown": self._df_to_markdown_clean(df),
                "metadata": {**metadata_base, "type": "excel_formulas", "count": len(rows)},
            }
        ]

    # ------------------------------------------------------------------
    # VBA macros
    # ------------------------------------------------------------------
    def _build_vba_documents(
        self, file_path: Path, metadata_base: Dict[str, Any]
    ) -> List[dict]:
        """Extract VBA macro code, if present, into markdown."""
        documents: List[dict] = []
        try:
            from oletools.olevba import VBA_Parser  # type: ignore
        except ImportError:
            # oletools is optional; skip if not installed
            self.logger.info(
                f"{file_path} - oletools not installed; skipping VBA macro extraction"
            )
            return documents

        try:
            vba = VBA_Parser(str(file_path))
        except Exception as exc:
            self.logger.error(f"{file_path} - Failed to open for VBA parsing: {exc}")
            return documents

        try:
            if not vba.detect_vba_macros():
                return documents

            snippets: List[str] = []
            for (_, _, vba_filename, vba_code) in vba.extract_macros():
                # basic guard against None
                code_str = vba_code or ""
                snippets.append(
                    f"### {vba_filename}\n\n```vba\n{code_str}\n```"
                )

            if not snippets:
                return documents

            documents.append(
                {
                    "markdown": "\n\n".join(snippets),
                    "metadata": {**metadata_base, "type": "excel_vba_macros"},
                }
            )
        except Exception as exc:
            self.logger.error(f"{file_path} - Error extracting VBA macros: {exc}")
        finally:
            try:
                vba.close()
            except Exception:
                pass

        return documents
