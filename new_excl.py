import openpyxl
from openpyxl.chart import (
    BarChart, LineChart, PieChart, AreaChart, ScatterChart,
    RadarChart, BubbleChart, DoughnutChart
)
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.chart.shapes import GraphicalProperties
from pathlib import Path
from typing import List, Dict, Optional, Tuple
import re
from io import BytesIO
from PIL import Image
import tempfile


class ExcelParser(BaseParser):
    def __init__(self) -> None:
        super().__init__(name="EXCEL_PARSER")
    
    def parse(self, file_path: Path, file_path_spaces: str) -> List[dict]:
        """Parse Excel file and extract all content including tables, charts, formulas, and macros."""
        documents = []
        
        try:
            self.logger.info(f"{file_path_spaces} - Loading Excel workbook...")
            # Load with data_only=True to get calculated values instead of formulas
            workbook = openpyxl.load_workbook(file_path, data_only=True, keep_vba=True)
            
            # Check for macros
            has_macros = workbook.vba_archive is not None
            if has_macros:
                self.logger.info(f"{file_path_spaces} - Workbook contains VBA macros")
            
            total_sheets = len(workbook.sheetnames)
            
            # Create temporary directory for images
            with tempfile.TemporaryDirectory(prefix="excel_img_render_") as tmp_img_dir:
                images_folder = Path(tmp_img_dir)
                
                for sheet_idx, sheet_name in enumerate(workbook.sheetnames):
                    self.logger.info(f"{file_path_spaces} - Processing sheet {sheet_idx + 1}/{total_sheets}: '{sheet_name}'")
                    
                    sheet = workbook[sheet_name]
                    
                    # Check if it's a Chartsheet (only contains charts) or regular Worksheet
                    if hasattr(sheet, 'iter_rows'):
                        # Regular worksheet
                        md_content = self.convert_sheet_to_markdown(
                            sheet, 
                            sheet_name, 
                            sheet_idx,
                            file_path,
                            file_path_spaces,
                            images_folder,
                            has_macros=has_macros
                        )
                    else:
                        # Chartsheet - only contains charts, no data cells
                        self.logger.info(f"{file_path_spaces} - Sheet '{sheet_name}' is a Chartsheet (chart-only)")
                        md_content = self.convert_chartsheet_to_markdown(
                            sheet,
                            sheet_name,
                            sheet_idx,
                            file_path,
                            file_path_spaces,
                            images_folder
                        )
                    
                    metadata = {
                        "file_path": file_path_spaces,
                        "sheet_name": sheet_name,
                        "sheet_number": sheet_idx + 1,
                        "has_macros": has_macros,
                        "file_hash": compute_file_hash(file_path)
                    }
                    metadata.update(self._load_sidecar_metadata(file_path))
                    
                    documents.append({
                        "markdown": md_content,
                        "metadata": metadata
                    })
            
            self.logger.success(f"✓ Converted {file_path_spaces}")
            return documents
            
        except Exception as e:
            self.logger.error(f"Failed to process {file_path_spaces}: {e}")
            return []
    
    def convert_sheet_to_markdown(self, sheet, sheet_name: str, sheet_idx: int, file_path: Path, 
                                  file_path_spaces: str, images_folder: Path, has_macros: bool = False) -> str:
        """Convert a single Excel sheet to markdown."""
        md_content = f"# Sheet: {sheet_name}\n\n"
        
        # Add macro warning if present
        if has_macros:
            md_content += "⚠️ **Note:** This workbook contains VBA macros.\n\n"
        
        # Check if sheet has charts first (before processing data)
        has_charts = hasattr(sheet, '_charts') and sheet._charts
        if has_charts:
            self.logger.info(f"{file_path_spaces} - Found {len(sheet._charts)} charts in sheet '{sheet_name}'")
        
        # Get the actual used range (excluding completely empty rows/columns)
        used_range = self._get_used_range(sheet)
        
        if not used_range:
            # Sheet has no data cells, but might have charts
            if has_charts:
                md_content += "*This sheet has no data cells, only charts.*\n\n"
                charts_md = self._extract_charts_with_vision(sheet, sheet_idx, file_path, file_path_spaces, images_folder)
                md_content += charts_md
            else:
                md_content += "*This sheet is empty.*\n\n"
            return md_content
        
        min_row, max_row, min_col, max_col = used_range
        self.logger.debug(f"{file_path_spaces} - Used range: rows {min_row}-{max_row}, cols {min_col}-{max_col}")
        
        # Extract tables with vision analysis
        tables_md = self._extract_tables_with_vision(
            sheet, 
            min_row, 
            max_row, 
            min_col, 
            max_col,
            sheet_idx,
            file_path,
            file_path_spaces,
            images_folder
        )
        md_content += tables_md
        
        # Extract charts with vision analysis (after tables for better organization)
        if has_charts:
            charts_md = self._extract_charts_with_vision(sheet, sheet_idx, file_path, file_path_spaces, images_folder)
            md_content += charts_md
        
        return md_content
    
    def convert_chartsheet_to_markdown(self, chartsheet, sheet_name: str, sheet_idx: int, 
                                      file_path: Path, file_path_spaces: str, images_folder: Path) -> str:
        """Convert a Chartsheet (chart-only sheet) to markdown."""
        md_content = f"# Chart Sheet: {sheet_name}\n\n"
        md_content += "*This is a dedicated chart sheet (contains only charts, no data cells).*\n\n"
        
        # Extract chart information if available
        if hasattr(chartsheet, '_charts') and chartsheet._charts:
            self.logger.info(f"{file_path_spaces} - Found {len(chartsheet._charts)} charts in chartsheet '{sheet_name}'")
            charts_md = self._extract_charts_with_vision(chartsheet, sheet_idx, file_path, file_path_spaces, images_folder)
            md_content += charts_md
        else:
            md_content += "*No charts found in this sheet.*\n\n"
        
        return md_content
    
    def _get_used_range(self, sheet) -> Optional[Tuple[int, int, int, int]]:
        """
        Get the actual used range of the sheet, excluding completely empty rows and columns.
        Returns: (min_row, max_row, min_col, max_col) or None if sheet is empty
        """
        min_row = None
        max_row = None
        min_col = None
        max_col = None
        
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None and str(cell.value).strip():
                    if min_row is None or cell.row < min_row:
                        min_row = cell.row
                    if max_row is None or cell.row > max_row:
                        max_row = cell.row
                    if min_col is None or cell.column < min_col:
                        min_col = cell.column
                    if max_col is None or cell.column > max_col:
                        max_col = cell.column
        
        if min_row is None:
            return None
        
        return (min_row, max_row, min_col, max_col)
    
    def _extract_tables_with_vision(
        self, 
        sheet, 
        min_row: int, 
        max_row: int, 
        min_col: int, 
        max_col: int,
        sheet_idx: int,
        file_path: Path,
        file_path_spaces: str,
        images_folder: Path
    ) -> str:
        """Extract tables and use GPT vision to describe them."""
        md_content = ""
        
        # Detect table regions (separated by blank rows/columns)
        table_regions = self._detect_table_regions(sheet, min_row, max_row, min_col, max_col)
        
        if not table_regions:
            # No distinct tables found, treat entire range as one table
            table_regions = [(min_row, max_row, min_col, max_col)]
        
        for table_idx, (t_min_row, t_max_row, t_min_col, t_max_col) in enumerate(table_regions):
            self.logger.debug(
                f"{file_path_spaces} - Processing table {table_idx + 1}: "
                f"rows {t_min_row}-{t_max_row}, cols {t_min_col}-{t_max_col}"
            )
            
            # Save table as image
            table_count = table_idx + 1
            img_path = Path(f"{images_folder}/{file_path.name.replace(' ', '-')}-sheet{sheet_idx}-table{table_count}.png")
            
            # Create a simple image representation of the table using PIL
            try:
                table_img = self._render_table_as_image(sheet, t_min_row, t_max_row, t_min_col, t_max_col)
                table_img.save(img_path)
                self.logger.debug(f"{file_path_spaces} - Table image saved at {img_path}")
                
                # Use GPT vision to describe the table
                self.logger.debug(f"{file_path_spaces} - Analyzing table {table_count} with GPT vision")
                table_description = self._md4vision(Path(img_path))
                
                md_content += f"\n**Table {table_count}**: {table_description}\n\n"
                
            except Exception as e:
                self.logger.error(f"{file_path_spaces} - Error processing table {table_count}: {e}")
                # Fallback: just note that a table exists
                md_content += f"\n**Table {table_count}**: [Table data present but could not be analyzed]\n\n"
        
        return md_content
    
    def _extract_charts_with_vision(self, sheet, sheet_idx: int, file_path: Path, 
                                    file_path_spaces: str, images_folder: Path) -> str:
        """Extract chart information using GPT vision."""
        if not hasattr(sheet, '_charts') or not sheet._charts:
            return ""
        
        md_content = "\n"
        
        for chart_idx, chart in enumerate(sheet._charts):
            chart_count = chart_idx + 1
            self.logger.debug(f"{file_path_spaces} - Processing chart {chart_count}")
            
            chart_title = chart.title if hasattr(chart, 'title') and chart.title else f"Chart {chart_count}"
            
            # Note: Excel charts through openpyxl don't easily export as images
            # We'll provide a description based on chart metadata and use vision if we can export
            try:
                # Get chart type
                chart_type = self._get_chart_type(chart)
                
                # Try to get chart data for context
                chart_data_summary = self._get_chart_data_summary(chart, sheet)
                
                # Create a text representation for vision analysis
                # In a real implementation, you'd export the chart as an image here
                # For now, we'll create a descriptive summary
                img_path = Path(f"{images_folder}/{file_path.name.replace(' ', '-')}-sheet{sheet_idx}-chart{chart_count}.png")
                
                # Try to render chart (this is a placeholder - actual chart rendering requires additional libraries)
                # In practice, you might use matplotlib or excel export functionality
                chart_description = f"A {chart_type}"
                if chart_data_summary:
                    chart_description += f" showing {chart_data_summary}"
                
                md_content += f"**Chart {chart_count} - {chart_title}**: {chart_description}\n\n"
                
            except Exception as e:
                self.logger.error(f"{file_path_spaces} - Error processing chart {chart_count}: {e}")
                md_content += f"**Chart {chart_count}**: [Chart present but could not be analyzed]\n\n"
        
        return md_content
    
    def _render_table_as_image(self, sheet, min_row: int, max_row: int, 
                               min_col: int, max_col: int) -> Image.Image:
        """Render a table region as an image for vision analysis."""
        from PIL import Image, ImageDraw, ImageFont
        
        # Calculate dimensions
        num_rows = max_row - min_row + 1
        num_cols = max_col - min_col + 1
        
        cell_width = 150
        cell_height = 30
        padding = 5
        
        img_width = num_cols * cell_width + padding * 2
        img_height = num_rows * cell_height + padding * 2
        
        # Create image
        img = Image.new('RGB', (img_width, img_height), color='white')
        draw = ImageDraw.Draw(img)
        
        # Try to use a font, fall back to default if not available
        try:
            font = ImageFont.truetype("arial.ttf", 12)
        except:
            font = ImageFont.load_default()
        
        # Draw table
        for row_idx, row in enumerate(range(min_row, max_row + 1)):
            for col_idx, col in enumerate(range(min_col, max_col + 1)):
                x = col_idx * cell_width + padding
                y = row_idx * cell_height + padding
                
                # Draw cell border
                draw.rectangle(
                    [x, y, x + cell_width, y + cell_height],
                    outline='black',
                    fill='lightgray' if row_idx == 0 else 'white'  # Header row highlighted
                )
                
                # Get cell value
                cell = sheet.cell(row, col)
                cell_value = self._format_cell_value(cell)
                
                # Draw text (truncate if too long)
                if len(cell_value) > 20:
                    cell_value = cell_value[:17] + "..."
                
                draw.text((x + 5, y + 8), cell_value, fill='black', font=font)
        
        return img
    
    def _get_chart_data_summary(self, chart, sheet) -> str:
        """Get a brief summary of chart data for context."""
        try:
            if not hasattr(chart, 'series') or not chart.series:
                return ""
            
            num_series = len(chart.series)
            series_names = []
            
            for series in chart.series[:3]:  # First 3 series
                if hasattr(series, 'title') and series.title:
                    series_names.append(str(series.title))
            
            if series_names:
                summary = f"{num_series} data series"
                if len(series_names) > 0:
                    summary += f" including {', '.join(series_names)}"
                return summary
            
            return f"{num_series} data series"
            
        except Exception:
            return ""
    
    def _detect_table_regions(
        self, 
        sheet, 
        min_row: int, 
        max_row: int, 
        min_col: int, 
        max_col: int
    ) -> List[Tuple[int, int, int, int]]:
        """
        Detect separate table regions within the sheet by identifying blank row/column separators.
        Returns list of (min_row, max_row, min_col, max_col) tuples.
        """
        # Simple implementation: split by blank rows
        # More sophisticated logic could split by blank columns too
        
        table_regions = []
        current_start_row = None
        
        for row_idx in range(min_row, max_row + 2):  # +2 to handle end of range
            # Check if row is blank
            row_is_blank = True
            if row_idx <= max_row:
                for col_idx in range(min_col, max_col + 1):
                    cell = sheet.cell(row_idx, col_idx)
                    if cell.value is not None and str(cell.value).strip():
                        row_is_blank = False
                        break
            
            if not row_is_blank and row_idx <= max_row:
                if current_start_row is None:
                    current_start_row = row_idx
            else:
                # Found blank row or end of range
                if current_start_row is not None:
                    table_regions.append((current_start_row, row_idx - 1, min_col, max_col))
                    current_start_row = None
        
        # Filter out very small regions (single row might be a label)
        table_regions = [r for r in table_regions if r[1] - r[0] >= 0]
        
        return table_regions
    
    def _convert_table_region_to_markdown(
        self, 
        sheet, 
        min_row: int, 
        max_row: int, 
        min_col: int, 
        max_col: int
    ) -> str:
        """
        This method is no longer used - tables are now converted to images
        and analyzed with vision. Kept for reference/backwards compatibility.
        """
        return ""
    
    def _format_cell_value(self, cell) -> str:
        """Format cell value for display (calculated values only, no formulas)."""
        if cell.value is None:
            return ""
        
        value = cell.value
        
        # Handle different data types
        if isinstance(value, (int, float)):
            # Check if it's formatted as percentage, currency, etc.
            if cell.number_format:
                if '%' in cell.number_format:
                    return f"{value:.1%}"
                elif '
    
    def _get_chart_type(self, chart) -> str:
        """Determine the chart type."""
        if isinstance(chart, BarChart):
            return "Bar Chart"
        elif isinstance(chart, LineChart):
            return "Line Chart"
        elif isinstance(chart, PieChart):
            return "Pie Chart"
        elif isinstance(chart, AreaChart):
            return "Area Chart"
        elif isinstance(chart, ScatterChart):
            return "Scatter Chart"
        elif isinstance(chart, RadarChart):
            return "Radar Chart"
        elif isinstance(chart, BubbleChart):
            return "Bubble Chart"
        elif isinstance(chart, DoughnutChart):
            return "Doughnut Chart"
        else:
            return "Chart"
 in cell.number_format or '€' in cell.number_format:
                    return f"${value:,.0f}"
            return str(value)
        
        # Clean up string values
        value_str = str(value).strip()
        
        # Escape pipe characters for markdown tables (only used in image rendering)
        value_str = value_str.replace('|', '\\|')
        
        # Preserve line breaks within cells
        value_str = value_str.replace('\n', '<br>')
        
        return value_str
    
    def _extract_charts(self, sheet, file_path_spaces: str) -> str:
        """Extract chart information from the sheet."""
        md_content = "\n## Charts\n\n"
        
        for chart_idx, chart in enumerate(sheet._charts):
            self.logger.debug(f"{file_path_spaces} - Processing chart {chart_idx + 1}")
            
            md_content += f"### Chart {chart_idx + 1}: {chart.title if hasattr(chart, 'title') and chart.title else 'Untitled'}\n\n"
            
            # Get chart type
            chart_type = self._get_chart_type(chart)
            md_content += f"**Type:** {chart_type}\n\n"
            
            # Extract chart data
            try:
                chart_data = self._extract_chart_data(chart, sheet)
                if chart_data:
                    md_content += chart_data + "\n"
            except Exception as e:
                self.logger.error(f"{file_path_spaces} - Error extracting chart data: {e}")
                md_content += "*Chart data could not be extracted.*\n\n"
        
        return md_content
    
    def _get_chart_type(self, chart) -> str:
        """Determine the chart type."""
        if isinstance(chart, BarChart):
            return "Bar Chart"
        elif isinstance(chart, LineChart):
            return "Line Chart"
        elif isinstance(chart, PieChart):
            return "Pie Chart"
        elif isinstance(chart, AreaChart):
            return "Area Chart"
        elif isinstance(chart, ScatterChart):
            return "Scatter Chart"
        elif isinstance(chart, RadarChart):
            return "Radar Chart"
        elif isinstance(chart, BubbleChart):
            return "Bubble Chart"
        elif isinstance(chart, DoughnutChart):
            return "Doughnut Chart"
        else:
            return "Unknown Chart Type"
    
    def _extract_chart_data(self, chart, sheet) -> str:
        """Extract data from chart series."""
        md_content = "**Data:**\n\n"
        
        if not hasattr(chart, 'series') or not chart.series:
            return "*No data series found.*\n"
        
        # Build table for chart data
        table_rows = []
        
        # Get categories (x-axis labels)
        categories = []
        if hasattr(chart, 'categories') and chart.categories:
            categories = self._parse_chart_reference(chart.categories, sheet)
        
        # Extract each series
        for series_idx, series in enumerate(chart.series):
            series_title = series.title if hasattr(series, 'title') and series.title else f"Series {series_idx + 1}"
            
            # Get series values
            if hasattr(series, 'val') and series.val:
                values = self._parse_chart_reference(series.val, sheet)
                
                # Create rows
                if not table_rows:
                    # Initialize with categories
                    if categories:
                        table_rows = [[cat] for cat in categories]
                    else:
                        table_rows = [[f"Point {i+1}"] for i in range(len(values))]
                
                # Add series values
                for i, value in enumerate(values):
                    if i < len(table_rows):
                        table_rows[i].append(str(value))
        
        if not table_rows:
            return "*No chart data available.*\n"
        
        # Build markdown table
        series_names = [s.title if hasattr(s, 'title') and s.title else f"Series {i+1}" 
                       for i, s in enumerate(chart.series)]
        
        header = ["Category"] + series_names
        md_lines = ["| " + " | ".join(header) + " |"]
        md_lines.append("| " + " | ".join(["---"] * len(header)) + " |")
        
        for row in table_rows:
            # Ensure row has correct number of columns
            while len(row) < len(header):
                row.append("")
            md_lines.append("| " + " | ".join(row[:len(header)]) + " |")
        
        md_content += "\n".join(md_lines) + "\n\n"
        
        return md_content
    
    def _parse_chart_reference(self, reference, sheet) -> List:
        """Parse chart data reference and extract values from sheet."""
        if not reference:
            return []
        
        # Handle different reference types
        if hasattr(reference, 'numRef') and reference.numRef:
            # Numerical reference
            if hasattr(reference.numRef, 'numCache') and reference.numRef.numCache:
                return [pt.v for pt in reference.numRef.numCache.pt]
        
        if hasattr(reference, 'strRef') and reference.strRef:
            # String reference
            if hasattr(reference.strRef, 'strCache') and reference.strRef.strCache:
                return [pt.v for pt in reference.strRef.strCache.pt]
        
        # Try to parse as cell range
        if hasattr(reference, 'f') and reference.f:
            formula = reference.f
            return self._extract_values_from_range(formula, sheet)
        
        return []
    
    def _extract_values_from_range(self, range_str: str, sheet) -> List:
        """Extract values from a cell range string like 'Sheet1!$A$1:$A$10'."""
        try:
            # Remove sheet name if present
            if '!' in range_str:
                range_str = range_str.split('!')[1]
            
            # Remove $ signs
            range_str = range_str.replace('$', '')
            
            # Parse range
            if ':' in range_str:
                start, end = range_str.split(':')
                start_col = column_index_from_string(re.sub(r'\d', '', start))
                start_row = int(re.sub(r'\D', '', start))
                end_col = column_index_from_string(re.sub(r'\d', '', end))
                end_row = int(re.sub(r'\D', '', end))
                
                values = []
                for row in range(start_row, end_row + 1):
                    for col in range(start_col, end_col + 1):
                        cell = sheet.cell(row, col)
                        if cell.value is not None:
                            values.append(cell.value)
                
                return values
        except Exception:
            return []
        
        return []
    
    def _extract_merged_cells_info(self, sheet) -> str:
        """Extract information about merged cells (indicates complex layouts)."""
        if not sheet.merged_cells.ranges:
            return ""
        
        md_content = "\n## Merged Cells\n\n"
        md_content += "*This sheet contains merged cells, which may indicate complex formatting:*\n\n"
        
        for merged_range in list(sheet.merged_cells.ranges)[:10]:  # Limit to first 10
            md_content += f"- {merged_range}\n"
        
        if len(sheet.merged_cells.ranges) > 10:
            md_content += f"- ... and {len(sheet.merged_cells.ranges) - 10} more\n"
        
        md_content += "\n"
        
        return md_content
